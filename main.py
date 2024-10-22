import os
import tkinter as tk
from tkinter import filedialog
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import json
import pandas as pd
from openpyxl import Workbook, load_workbook
import threading


class App:
    def __init__(self, master):
        self.master = master
        self.master.title("公司地址获取")

        self.master.geometry("800x600")

        # 输入文件选择
        self.input_button = tk.Button(master, text="选择输入文件", command=self.select_input_file)
        self.input_button.pack(pady=10)

        # 输出文件选择
        self.output_button = tk.Button(master, text="选择输出文件", command=self.select_output_file)
        self.output_button.pack(pady=10)

        # 开始按钮
        self.start_button = tk.Button(master, text="开始获取地址", command=self.start_process)
        self.start_button.pack(pady=10)

        # 停止按钮
        self.stop_button = tk.Button(master, text="停止", command=self.stop_process, state=tk.DISABLED)
        self.stop_button.pack(pady=10)

        # 状态标签
        self.status_label = tk.Label(master, text="")
        self.status_label.pack(pady=10)

        self.is_running = False
        self.worker_thread = None

    def select_input_file(self):
        self.input_file = filedialog.askopenfilename(title="选择输入的 Excel 文件",
                                                     filetypes=[("Excel 文件", "*.xlsx")])
        if self.input_file:
            self.status_label.config(text=f"已选择输入文件: {self.input_file}")

    def select_output_file(self):
        self.output_file = filedialog.asksaveasfilename(title="选择输出的 Excel 文件名", defaultextension=".xlsx",
                                                        filetypes=[("Excel 文件", "*.xlsx")])
        if self.output_file:
            self.status_label.config(text=f"已选择输出文件: {self.output_file}")

    def start_process(self):
        if not hasattr(self, 'input_file') or not hasattr(self, 'output_file'):
            self.status_label.config(text="请先选择输入和输出文件。")
            return

        self.is_running = True
        self.stop_button.config(state=tk.NORMAL)
        self.start_button.config(state=tk.DISABLED)
        self.status_label.config(text="正在获取地址，请稍候...")

        self.worker_thread = threading.Thread(target=self.process_addresses)
        self.worker_thread.start()

    def stop_process(self):
        self.is_running = False
        self.status_label.config(text="程序已停止。")
        self.start_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)

    def process_addresses(self):
        try:
            # Excel 文件初始化
            try:
                self.processed_count = 0
                workbook = load_workbook(self.output_file)
                sheet = workbook.active
                existing_companies = {row[0].value for row in sheet.iter_rows(min_row=2, values_only=True) if row[0]}
            except FileNotFoundError:
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(['企业名称', '地址'])
                existing_companies = set()

            # 读取输入的 Excel 文件，获取公司名称列表
            df = pd.read_excel(self.input_file)
            list_name = df['企业名称'].tolist()

            # 获取每个公司的地址，并将其增量写入 Excel 文件
            for name in list_name:
                if not self.is_running:
                    break  # 如果被停止，终止处理
                if name in existing_companies:
                    print(f"Skipping {name}, already exists in the file.")
                    continue

                address = self.get_address(name)
                sheet.append([name, address])
                existing_companies.add(name)
                workbook.save(self.output_file)
                print(f"Saved address for {name}: {address}")

                # 更新计数器
                self.processed_count += 1
                # 更新状态标签
                self.status_label.config(text=f"正在获取地址，已处理 {self.processed_count} 个公司...")

            workbook.save(self.output_file)
            workbook.close()
            self.status_label.config(text="地址获取完成。")
        except Exception as e:
            self.status_label.config(text=f"发生错误: {e}")
        finally:
            self.start_button.config(state=tk.NORMAL)
            self.stop_button.config(state=tk.DISABLED)

    def get_address(self, name):
        driver = webdriver.Chrome()
        address = "未获取到"  # 默认值
        try:
            # driver.get('https://www.tianyancha.com/nadvance')
            driver.get('https://www.tianyancha.com')
            driver.maximize_window()
            driver.delete_all_cookies()

            # 加载 cookies 文件
            with open('cookies.txt', 'r') as f:
                cookies_list = json.load(f)
                for cookie in cookies_list:
                    if isinstance(cookie.get('expiry'), float):
                        cookie['expiry'] = int(cookie['expiry'])
                    driver.add_cookie(cookie)

            driver.refresh()
            driver.implicitly_wait(100)

            # search_box = driver.find_element(By.ID, "header-company-search")
            # search_box.click()
            # search_box.send_keys(name)
            driver.find_element(By.CSS_SELECTOR, ".\\_e096d").send_keys(name)
            driver.find_element(By.CSS_SELECTOR, ".index_home-suggest-button__GuWyT > span").click()
            # driver.find_element(By.CSS_SELECTOR, ".component").click()

            # 等待地址元素出现
            address_element = WebDriverWait(driver, 100).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".index_address__mHjQD > .index_value__Pl0Nh"))
            )
            address = address_element.text
        except Exception as e:
            print(f"Error retrieving address for {name}: {e}")
        finally:
            driver.quit()

        return address


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
